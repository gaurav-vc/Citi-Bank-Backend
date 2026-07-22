import csv
import io
from openpyxl import load_workbook

def normalize_header(h):
    if h is None:
        return ""
    h = str(h).strip().lower().replace(' ', '_').replace('-', '_').replace('/', '_')
    aliases = {
        'item_id': 'id',
        'item_code': 'id',
        'code': 'id',
        'current_stock_level': 'current_stock',
        'stock': 'current_stock',
        'qty': 'current_stock',
        'quantity': 'current_stock',
        'rate': 'unit_price',
        'price': 'unit_price',
        'unit_rate': 'unit_price',
    }
    return aliases.get(h, h)

def parse_csv(file_content):
    try:
        decoded_file = file_content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            decoded_file = file_content.decode('latin-1')
        except Exception as e:
            raise ValueError(f"Could not decode CSV file: {e}")
            
    csv_file = io.StringIO(decoded_file)
    reader = csv.reader(csv_file)
    
    headers = next(reader, None)
    if not headers:
        return []
    
    normalized_headers = [normalize_header(h) for h in headers]
    rows = []
    for row in reader:
        if not row or all(cell == '' for cell in row):
            continue
        row_dict = {}
        for idx, val in enumerate(row):
            if idx < len(normalized_headers):
                row_dict[normalized_headers[idx]] = val.strip()
        rows.append(row_dict)
    return rows

def parse_excel(file_content):
    file_bytes = io.BytesIO(file_content)
    try:
        wb = load_workbook(file_bytes, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")
        
    ws = wb.active
    
    rows = []
    headers = None
    for r in ws.iter_rows(values_only=True):
        # Skip empty rows before header
        if not headers:
            if any(cell is not None for cell in r):
                headers = [normalize_header(cell) for cell in r]
            continue
        
        # Check if row is empty
        if all(cell is None or str(cell).strip() == '' for cell in r):
            continue
            
        row_dict = {}
        for idx, val in enumerate(r):
            if idx < len(headers):
                header_name = headers[idx]
                if header_name:
                    cell_val = str(val).strip() if val is not None else ''
                    row_dict[header_name] = cell_val
        rows.append(row_dict)
    return rows

def parse_uploaded_file(uploaded_file):
    filename = uploaded_file.name.lower()
    file_content = uploaded_file.read()
    
    if filename.endswith('.csv'):
        return parse_csv(file_content)
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        return parse_excel(file_content)
    else:
        raise ValueError("Unsupported file format. Please upload CSV or Excel.")
