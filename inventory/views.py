from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from access_control.permissions import RBACPermission
from rest_framework.decorators import action
from django.db import transaction, connection
from django.http import HttpResponse, FileResponse
from datetime import datetime, date
from io import BytesIO
import openpyxl

from .models import Item, GRN, StockTransfer, MaterialIssue, ScrapDisposal
from .serializers import (
    ItemSerializer, GRNSerializer, StockTransferSerializer,
    MaterialIssueSerializer, ScrapDisposalSerializer
)
from utils.exporter import export_data
from utils.importer import parse_uploaded_file
from utils.db_logger import log_export, log_import_start, log_import_failed_row, log_import_end

class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated, RBACPermission]

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        try:
            format = request.query_params.get('format', 'xlsx')
            items = Item.objects.all().order_by('id')
            data = [{
                'Item ID': i.id,
                'Item Name': i.name,
                'Type': i.type,
                'Category': i.category,
                'UOM': i.uom,
                'Min Stock Level': i.min_stock_level,
                'Reorder Level': i.reorder_level,
                'Current Stock': i.current_stock,
                'Preferred Vendor': i.preferred_vendor or '',
                'Unit Price': float(i.unit_price)
            } for i in items]
            
            columns = [
                {'header': 'Item ID', 'key': 'Item ID'},
                {'header': 'Item Name', 'key': 'Item Name'},
                {'header': 'Type', 'key': 'Type'},
                {'header': 'Category', 'key': 'Category'},
                {'header': 'UOM', 'key': 'UOM'},
                {'header': 'Min Stock Level', 'key': 'Min Stock Level'},
                {'header': 'Reorder Level', 'key': 'Reorder Level'},
                {'header': 'Current Stock', 'key': 'Current Stock'},
                {'header': 'Preferred Vendor', 'key': 'Preferred Vendor'},
                {'header': 'Unit Price', 'key': 'Unit Price'}
            ]
            
            log_export('items', f"items_export_{int(datetime.now().timestamp())}.xlsx", 'xlsx', request.query_params)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Items"
            ws.append([col['header'] for col in columns])
            for row in data:
                ws.append([row.get(col['key'], '') for col in columns])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return FileResponse(
                buffer,
                as_attachment=True,
                filename=f"items_export_{int(datetime.now().timestamp())}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='import')
    def import_items(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            rows = parse_uploaded_file(uploaded_file)
            imported_count = 0
            errors = []
            
            for idx, row in enumerate(rows):
                row_idx = idx + 2
                try:
                    item_id = row.get('id')
                    if not item_id:
                        errors.append(f"Row {row_idx}: Missing Item ID/Code")
                        continue
                    
                    name = row.get('name') or row.get('item_name') or ''
                    item_type = row.get('type') or row.get('item_type') or 'spare'
                    category = row.get('category') or ''
                    uom = row.get('uom') or 'Nos'
                    
                    try:
                        min_stock = int(float(row.get('min_stock_level') or 0))
                    except ValueError:
                        min_stock = 0
                        
                    try:
                        reorder = int(float(row.get('reorder_level') or 0))
                    except ValueError:
                        reorder = 0
                        
                    try:
                        current_stock = int(float(row.get('current_stock') or 0))
                    except ValueError:
                        current_stock = 0
                        
                    preferred_vendor = row.get('preferred_vendor') or ''
                    
                    try:
                        unit_price_val = row.get('unit_price') or 0
                        unit_price = float(unit_price_val)
                    except ValueError:
                        unit_price = 0.0

                    Item.objects.update_or_create(
                        id=item_id,
                        defaults={
                            'name': name,
                            'type': item_type,
                            'category': category,
                            'uom': uom,
                            'min_stock_level': min_stock,
                            'reorder_level': reorder,
                            'current_stock': current_stock,
                            'preferred_vendor': preferred_vendor,
                            'unit_price': unit_price
                        }
                    )
                    imported_count += 1
                except Exception as row_err:
                    errors.append(f"Row {row_idx}: {str(row_err)}")
            
            return Response({"imported": imported_count, "errors": errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

from .serializers import ProductInspectionSerializer
from .models import ProductInspection
from django.utils import timezone

class ProductInspectionViewSet(viewsets.ModelViewSet):
    queryset = ProductInspection.objects.all().order_by('-created_at')
    serializer_class = ProductInspectionSerializer
    permission_classes = [IsAuthenticated, RBACPermission]

    def perform_create(self, serializer):
        po_id = self.request.data.get('po_id')
        vendor_name = ""
        if po_id:
            from procurement.models import PurchaseOrder
            po = PurchaseOrder.objects.filter(id=po_id).first()
            if po:
                vendor_name = po.vendor_name
        serializer.save(vendor_name=vendor_name, status='pending')

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        new_status = request.data.get('status', instance.status)
        if new_status == 'completed' and instance.status != 'completed':
            # Create GRN for accepted items
            items = request.data.get('items', instance.items)
            accepted_items = []
            for item in items:
                # support both camelCase and snake_case for frontend
                acc = int(item.get('acceptedQty', item.get('accepted_qty', 0)))
                if acc > 0:
                    accepted_items.append(item)
            
            if accepted_items:
                grn_id = f"GRN-{int(timezone.now().timestamp())}"
                GRN.objects.create(
                    id=grn_id,
                    po_id=instance.po_id,
                    vendor_name=instance.vendor_name,
                    received_date=instance.received_date,
                    received_by=instance.inspector_name,
                    items=accepted_items,
                    status='pending', # GRN is created but maybe needs further approval or just accepted
                    invoice_number=instance.invoice_number,
                    invoice_date=instance.invoice_date,
                    remarks=instance.remarks,
                    attachments=instance.attachments
                )
        return super().update(request, *args, **kwargs)

class GRNViewSet(viewsets.ModelViewSet):
    queryset = GRN.objects.all().order_by('-created_at')
    serializer_class = GRNSerializer
    permission_classes = [IsAuthenticated, RBACPermission]

    def create(self, request, *args, **kwargs):
        invoice_number = request.data.get('invoice_number')
        if invoice_number:
            from procurement.models import Invoice
            if Invoice.objects.filter(invoice_number=invoice_number).exists() or GRN.objects.filter(invoice_number=invoice_number).exists():
                return Response(
                    {"error": "Invoice number already exists."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        invoice_number = request.data.get('invoice_number')
        instance = self.get_object()
        if invoice_number and invoice_number != instance.invoice_number:
            from procurement.models import Invoice
            if Invoice.objects.filter(invoice_number=invoice_number).exists() or GRN.objects.filter(invoice_number=invoice_number).exists():
                return Response(
                    {"error": "Invoice number already exists."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        return super().update(request, *args, **kwargs)

    def perform_create(self, serializer):
        po_id = self.request.data.get('po_id')
        vendor_name = ""
        if po_id:
            from procurement.models import PurchaseOrder
            po = PurchaseOrder.objects.filter(id=po_id).first()
            if po:
                vendor_name = po.vendor_name
        serializer.save(vendor_name=vendor_name, status='pending_qc')

    def retrieve(self, request, *args, **kwargs):
        from procurement.models import PurchaseOrder
        from vendors.models import Vendor

        grn = self.get_object()
        po = PurchaseOrder.objects.filter(id=grn.po_id).first()
        
        vendor_data = {
            "id": "",
            "name": grn.vendor_name or "System Vendor"
        }
        purchase_order_data = None
        attachments = []

        if po:
            vendor_obj = Vendor.objects.filter(id=po.vendor).first()
            if vendor_obj:
                vendor_data = {
                    "id": vendor_obj.id,
                    "name": vendor_obj.name
                }
            else:
                vendor_data = {
                    "id": po.vendor,
                    "name": po.vendor_name
                }
            purchase_order_data = {
                "id": po.id,
                "po_number": po.id,
                "po_date": po.start_date.isoformat() if po.start_date else "",
                "department": po.category or "",
                "requester": po.created_by or "",
                "expected_delivery_date": po.end_date.isoformat() if po.end_date else ""
            }
            if isinstance(po.attachments, list):
                attachments = po.attachments

        formatted_items = []
        for item in grn.items:
            item_name = item.get('item_name') or item.get('itemName') or 'Unknown Item'
            ordered_qty = item.get('ordered_qty') or item.get('orderedQty') or 0
            received_qty = item.get('received_qty') or item.get('receivedQty') or 0
            accepted_qty = item.get('accepted_qty') or item.get('acceptedQty') or 0
            rejected_qty = item.get('rejected_qty') or item.get('rejectedQty') or 0
            
            # Find unit price
            unit_price = 0.0
            if po and isinstance(po.items, list):
                po_item = next((pi for pi in po.items if pi.get('itemName') == item_name or pi.get('itemId') == item.get('item_id') or pi.get('itemId') == item.get('itemId')), None)
                if po_item:
                    unit_price = float(po_item.get('rate') or po_item.get('unit_price') or 0.0)
            
            if not unit_price:
                item_obj = Item.objects.filter(name=item_name).first()
                if item_obj:
                    unit_price = float(item_obj.unit_price)

            # Quality Inspection section requirements
            inspection_status = "Pending"
            if accepted_qty > 0 and rejected_qty == 0:
                inspection_status = "Accepted"
            elif rejected_qty > 0 and accepted_qty == 0:
                inspection_status = "Rejected"
            elif accepted_qty > 0 and rejected_qty > 0:
                inspection_status = "Partially Accepted"

            formatted_items.append({
                "item_name": item_name,
                "ordered_quantity": ordered_qty,
                "received_quantity": received_qty,
                "accepted_quantity": accepted_qty,
                "rejected_quantity": rejected_qty,
                "unit_price": unit_price,
                "inspection_status": inspection_status,
                "remarks": item.get('remarks') or grn.remarks or '',
                "rejection_reason": item.get('rejection_reason') or (item.get('remarks') if rejected_qty > 0 else '')
            })

        return Response({
            "id": grn.id,
            "grn_number": grn.id,
            "po_reference": grn.po_id,
            "vendor": vendor_data,
            "invoice_number": grn.invoice_number or "",
            "invoice_date": grn.invoice_date.isoformat() if grn.invoice_date else "",
            "received_date": grn.received_date.isoformat() if grn.received_date else "",
            "received_by": grn.received_by,
            "status": grn.status.capitalize() if grn.status else "Pending",
            "created_date": grn.created_at.isoformat() if grn.created_at else "",
            "last_updated_date": grn.created_at.isoformat() if grn.created_at else "",
            "purchase_order": purchase_order_data,
            "items": formatted_items,
            "attachments": grn.attachments or []
        })

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        try:
            format = request.query_params.get('format', 'xlsx')
            grns = GRN.objects.all().order_by('-created_at')
            data = [{
                'id': g.id,
                'po_id': g.po_id,
                'vendor_name': g.vendor_name,
                'invoice_number': g.invoice_number,
                'received_date': g.received_date.isoformat() if isinstance(g.received_date, (date, datetime)) else str(g.received_date),
                'received_by': g.received_by,
                'status': g.status,
                'remarks': g.remarks,
            } for g in grns]

            columns = [
                {'header': 'GRN ID', 'key': 'id'},
                {'header': 'PO ID', 'key': 'po_id'},
                {'header': 'Vendor Name', 'key': 'vendor_name'},
                {'header': 'Invoice Number', 'key': 'invoice_number'},
                {'header': 'Received Date', 'key': 'received_date'},
                {'header': 'Received By', 'key': 'received_by'},
                {'header': 'Status', 'key': 'status'},
                {'header': 'Remarks', 'key': 'remarks'},
            ]

            log_export('grns', f"grns_export_{int(datetime.now().timestamp())}.xlsx", 'xlsx', request.query_params)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GRNs"
            ws.append([col['header'] for col in columns])
            for row in data:
                ws.append([row.get(col['key'], '') for col in columns])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return FileResponse(
                buffer,
                as_attachment=True,
                filename=f"grns_export_{int(datetime.now().timestamp())}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='action')
    @transaction.atomic
    def process_action(self, request, pk=None):
        grn = self.get_object()
        
        # QC Acceptance Rule: Block if inspection is not completed
        if grn.status != 'qc_completed':
            return Response(
                {"error": "Please complete Quality Inspection before processing this GRN."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Determine Final GRN Status based on inspection data saved in grn.items
        total_received = sum(int(item.get('received_qty', item.get('receivedQty', 0))) for item in grn.items)
        total_accepted = sum(int(item.get('accepted_qty', item.get('acceptedQty', 0))) for item in grn.items)
        total_rejected = sum(int(item.get('rejected_qty', item.get('rejectedQty', 0))) for item in grn.items)

        if total_accepted == total_received and total_rejected == 0:
            action_type = 'accept'
        elif total_accepted == 0 and total_rejected == total_received:
            action_type = 'reject'
        else:
            action_type = 'partial_accept'

        inventory_decision = request.data.get('inventory_decision', 'surplus') # 'surplus' or 'site'

        if action_type in ['accept', 'partial_accept']:
            invoice_num = grn.invoice_number or f"SYS-{grn.id}"
            from procurement.models import Invoice
            if Invoice.objects.filter(invoice_number=invoice_num).exists():
                return Response(
                    {"error": "Invoice number already exists."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Initialize GDN and RTV lists
        gdn_items = []
        rtv_items = []

        if action_type == 'reject':
            grn.status = 'rejected'
            grn.save()
            from .models import ReturnToVendor, StockLedger
            import time
            import uuid
            rtv_id = f"RTV-{int(time.time())}"
            ReturnToVendor.objects.create(
                id=rtv_id,
                grn_id=grn.id,
                vendor_reference=grn.vendor_name,
                items=grn.items,
                reason="GRN Rejected entirely",
                status='returned'
            )
            for item in grn.items:
                item_name = item.get('itemName') or item.get('item_name')
                inv_item = Item.objects.filter(name=item_name).first()
                if inv_item:
                    qty = int(item.get('receivedQty', item.get('received_qty', item.get('orderedQty', 0))))
                    StockLedger.objects.create(
                        id=f'SL-{uuid.uuid4().hex[:8].upper()}',
                        transaction_type='RETURN_TO_VENDOR',
                        source_type='RTV',
                        source_id=rtv_id,
                        item=inv_item,
                        quantity=qty,
                        stock_before=inv_item.current_stock,
                        stock_after=inv_item.current_stock,
                        user=request.user if request.user.is_authenticated else None,
                        reason="GRN Rejected entirely",
                        reference_number=grn.vendor_name,
                        remarks="Full rejection"
                    )
            return Response({'status': 'GRN rejected, Return-To-Vendor initiated'})

        if action_type in ['accept', 'partial_accept']:
            grn.status = 'accepted' if action_type == 'accept' else 'partial_accepted'
            
            # Quantities and remarks are already stored in grn.items during the inspection view.
            # Do NOT update grn.items here with items_data as it was verified during QC.
            grn.save()

            # Update inventory stock and calculate invoice total
            import logging
            logger = logging.getLogger(__name__)
            from procurement.models import PurchaseOrder
            
            po = PurchaseOrder.objects.filter(id=grn.po_id).first()
            invoice_total = 0
            
            for item in grn.items:
                accepted_qty = int(item.get('accepted_qty', item.get('acceptedQty', 0)))
                rejected_qty = int(item.get('rejected_qty', item.get('rejectedQty', 0)))
                item_name = item.get('itemName') or item.get('item_name', 'Unknown Item')
                
                if rejected_qty > 0:
                    rtv_items.append({
                        'itemName': item_name,
                        'quantity': rejected_qty,
                        'reason': item.get('remarks', 'Rejected during Partial Accept')
                    })
                    try:
                        inv_item = Item.objects.filter(name=item_name).first()
                        if inv_item:
                            from .models import StockLedger
                            import uuid
                            StockLedger.objects.create(
                                id=f'SL-{uuid.uuid4().hex[:8].upper()}',
                                transaction_type='RETURN_TO_VENDOR',
                                source_type='RTV',
                                source_id=f"RTV-{grn.id}",
                                item=inv_item,
                                quantity=rejected_qty,
                                stock_before=inv_item.current_stock,
                                stock_after=inv_item.current_stock,
                                user=request.user if request.user.is_authenticated else None,
                                reason="Partial Rejection",
                                reference_number=grn.vendor_name,
                                remarks=item.get('remarks', 'Rejected during Partial Accept')
                            )
                    except Exception as e:
                        logger.error(f"Error creating RTV ledger for {item_name}: {e}")

                if accepted_qty > 0:
                    if inventory_decision == 'site':
                        gdn_items.append({
                            'itemName': item_name,
                            'quantity': accepted_qty,
                            'status': 'dispatched'
                        })
                    
                    try:
                        inv_item = Item.objects.filter(name=item_name).first()
                        if inv_item:
                            before_stock = inv_item.current_stock
                            
                            from .models import StockLedger
                            import uuid
                            StockLedger.objects.create(
                                id=f'SL-{uuid.uuid4().hex[:8].upper()}',
                                transaction_type='GRN_RECEIPT',
                                source_type='GRN',
                                source_id=grn.id,
                                item=inv_item,
                                quantity=accepted_qty,
                                stock_before=before_stock,
                                stock_after=before_stock + accepted_qty,
                                user=request.user if request.user.is_authenticated else None,
                                reason="GRN Acceptance",
                                reference_number=grn.po_id,
                                remarks=f"Received via PO {grn.po_id}"
                            )
                            
                            inv_item.current_stock += accepted_qty
                            
                            if inventory_decision == 'site':
                                StockLedger.objects.create(
                                    id=f'SL-{uuid.uuid4().hex[:8].upper()}',
                                    transaction_type='GOODS_ISSUE',
                                    source_type='GDN',
                                    source_id=f"GDN-{grn.id}",
                                    item=inv_item,
                                    quantity=accepted_qty,
                                    stock_before=inv_item.current_stock,
                                    stock_after=inv_item.current_stock - accepted_qty,
                                    user=request.user if request.user.is_authenticated else None,
                                    reason="Site Issue / GDN",
                                    reference_number=grn.po_id,
                                    remarks="Issued directly to site"
                                )
                                inv_item.current_stock -= accepted_qty
                                
                            inv_item.save()
                            
                            logger.info(f"Inventory Update - Item: {item_name}, Before: {before_stock}, Accepted Qty: {accepted_qty}, After: {inv_item.current_stock}")
                            
                            po_rate = float(inv_item.unit_price) # Fallback
                            if po and isinstance(po.items, list):
                                po_line = next((i for i in po.items if i.get('itemId') == (item.get('itemId') or item.get('item_id')) or i.get('itemName') == item_name), None)
                                if po_line:
                                    po_rate = float(po_line.get('rate', po_line.get('unit_price', po_rate)))
                                    
                            invoice_total += po_rate * accepted_qty
                    except Exception as e:
                        logger.error(f"Error updating inventory for {item_name}: {e}")

            if inventory_decision == 'site' and gdn_items:
                from .models import GoodsDispatchNote
                import time
                GoodsDispatchNote.objects.create(
                    id=f"GDN-{int(time.time())}",
                    grn_id=grn.id,
                    destination="Project Site",
                    items=gdn_items,
                    status="dispatched"
                )

            if rtv_items:
                from .models import ReturnToVendor
                import time
                ReturnToVendor.objects.create(
                    id=f"RTV-{int(time.time())}",
                    grn_id=grn.id,
                    vendor_reference=grn.vendor_name,
                    items=rtv_items,
                    reason="Partial Rejection",
                    status="returned"
                )

            # Auto-generate Invoice and Approval Request for Finance Executive
            from procurement.models import Invoice, PurchaseOrder
            from approvals.models import ApprovalRequest
            from django.contrib.auth import get_user_model
            import time

            po = PurchaseOrder.objects.filter(id=grn.po_id).first()
            actual_vendor_id = po.vendor if po else grn.vendor_name

            User = get_user_model()
            finance_execs = User.objects.filter(role='finance_executive')
            assigned_user = finance_execs.first() if finance_execs.exists() else request.user

            invoice = Invoice.objects.create(
                id=f"INV-{int(time.time())}",
                vendor_id=actual_vendor_id,
                vendor_name=grn.vendor_name,
                invoice_number=grn.invoice_number or f"SYS-{grn.id}",
                invoice_date=grn.invoice_date or date.today(),
                po_id=grn.po_id,
                grn_id=grn.id,
                amount=invoice_total,
                gst=invoice_total * 0.18, # Default 18% GST for simulation
                total_amount=invoice_total * 1.18,
                due_date=date.today(),
                status='pending_approval',
                remarks=grn.remarks,
                attachments=grn.attachments or []
            )

            ApprovalRequest.objects.create(
                entity_type='invoice',
                entity_id=invoice.id,
                requested_by=request.user,
                assigned_to=assigned_user,
                status='pending',
                remarks=f"Auto-generated for GRN {grn.id}"
            )

            return Response({'status': f'GRN {action_type}ed, Invoice generated and routed to Finance'})

class StockTransferViewSet(viewsets.ModelViewSet):
    queryset = StockTransfer.objects.all()
    serializer_class = StockTransferSerializer
    permission_classes = [IsAuthenticated, RBACPermission]

class MaterialIssueViewSet(viewsets.ModelViewSet):
    queryset = MaterialIssue.objects.all()
    serializer_class = MaterialIssueSerializer
    permission_classes = [IsAuthenticated, RBACPermission]

class ScrapDisposalViewSet(viewsets.ModelViewSet):
    queryset = ScrapDisposal.objects.all()
    serializer_class = ScrapDisposalSerializer
    permission_classes = [IsAuthenticated, RBACPermission]

# Custom views mapped to frontend /api/inventory/* endpoints
class StockLedgerView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            items = Item.objects.all().order_by('id')
            mapped = [{
                'id': i.id,
                'name': i.name,
                'type': i.type,
                'category': i.category,
                'uom': i.uom,
                'minStockLevel': i.min_stock_level,
                'reorderLevel': i.reorder_level,
                'currentStock': i.current_stock,
                'preferredVendor': i.preferred_vendor or '',
                'unitPrice': float(i.unit_price)
            } for i in items]
            return Response(mapped)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class InventoryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            format = request.query_params.get('format', 'xlsx')
            items = Item.objects.all().order_by('id')
            data = [{
                'id': i.id,
                'name': i.name,
                'category': i.category,
                'type': i.type,
                'current_stock': i.current_stock,
                'uom': i.uom,
                'min_stock_level': i.min_stock_level,
                'reorder_level': i.reorder_level,
                'unit_price': float(i.unit_price)
            } for i in items]
            
            columns = [
                {'header': 'Item ID', 'key': 'id'},
                {'header': 'Item Name', 'key': 'name'},
                {'header': 'Category', 'key': 'category'},
                {'header': 'Type', 'key': 'type'},
                {'header': 'Current Stock', 'key': 'current_stock'},
                {'header': 'UOM', 'key': 'uom'},
                {'header': 'Min Stock Level', 'key': 'min_stock_level'},
                {'header': 'Reorder Level', 'key': 'reorder_level'},
                {'header': 'Unit Price', 'key': 'unit_price'}
            ]
            
            log_export('inventory', f"inventory_export_{int(datetime.now().timestamp())}.xlsx", 'xlsx', request.query_params)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory Stock"
            ws.append([col['header'] for col in columns])
            for row in data:
                ws.append([row.get(col['key'], '') for col in columns])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return FileResponse(
                buffer,
                as_attachment=True,
                filename=f"inventory_export_{int(datetime.now().timestamp())}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class InventoryImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            rows = parse_uploaded_file(uploaded_file)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        file_type = uploaded_file.name.split('.')[-1].lower()
        log_id = log_import_start('inventory', uploaded_file.name, file_type, len(rows))
        
        validation_errors = []
        valid_rows = []
        
        for idx, row in enumerate(rows):
            row_idx = idx + 1
            row_errors = []
            
            item_id = row.get('id')
            if not item_id:
                row_errors.append("Missing Item ID")
            
            stock_val = row.get('current_stock')
            if stock_val == '' or stock_val is None:
                row_errors.append("Missing Current Stock")
            else:
                try:
                    int(float(stock_val))
                except ValueError:
                    row_errors.append("Current Stock must be an integer")
                    
            if row_errors:
                err_msg = "; ".join(row_errors)
                validation_errors.append(f"Row {row_idx}: {err_msg}")
                log_import_failed_row(log_id, row_idx, row, err_msg)
            else:
                valid_rows.append((row, row_idx))
                
        if validation_errors:
            log_import_end(log_id, 'failed', 0, len(rows))
            return Response({
                'error': 'Import failed validation',
                'errors': validation_errors,
                'totalRows': len(rows),
                'processedRows': 0,
                'failedRows': len(rows)
            }, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
            
        try:
            with transaction.atomic():
                for row, row_idx in valid_rows:
                    item_id = row['id']
                    stock_qty = int(float(row['current_stock']))
                    
                    item = Item.objects.filter(id=item_id).first()
                    if not item:
                        raise ValueError(f"Row {row_idx}: Item code \"{item_id}\" does not exist in items master")
                        
                    item.current_stock = stock_qty
                    item.save()
                    
            log_import_end(log_id, 'success', len(valid_rows), 0)
            return Response({
                'success': True,
                'message': f"Successfully imported {len(valid_rows)} stock level updates",
                'totalRows': len(rows),
                'processedRows': len(valid_rows),
                'failedRows': 0
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            log_import_end(log_id, 'failed', 0, len(rows))
            return Response({'error': f"Stock update error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from .models import StockLedger
from vendors.models import AuditLog
import uuid

class AddStockView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        user = request.user
        if getattr(user, 'role', '') not in ['store_keeper', 'super_admin']:
            return Response({'error': 'Permission denied. Only store_keeper or super_admin can manually add stock.'}, status=status.HTTP_403_FORBIDDEN)
        
        item_id = request.data.get('item_id')
        quantity = request.data.get('quantity')
        reason = request.data.get('reason')
        reference_number = request.data.get('reference_number', '')
        remarks = request.data.get('remarks', '')

        if not all([item_id, quantity, reason, remarks]):
            return Response({'error': 'Missing required fields: item_id, quantity, reason, remarks'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            quantity = int(quantity)
            if quantity <= 0:
                return Response({'error': 'Quantity must be strictly > 0'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({'error': 'Quantity must be a number'}, status=status.HTTP_400_BAD_REQUEST)
            
        if len(remarks) < 10:
            return Response({'error': 'Remarks must be at least 10 characters long'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            item = Item.objects.get(id=item_id)
        except Item.DoesNotExist:
            return Response({'error': 'Item not found'}, status=status.HTTP_404_NOT_FOUND)

        stock_before = item.current_stock
        item.current_stock += quantity
        item.save()

        import time
        StockLedger.objects.create(
            id=f'SL-{uuid.uuid4().hex[:8].upper()}',
            transaction_type='MANUAL_ADJUSTMENT',
            source_type='MANUAL',
            source_id=f"MANUAL-{int(time.time())}",
            item=item,
            quantity=quantity,
            stock_before=stock_before,
            stock_after=item.current_stock,
            user=user if user.is_authenticated else None,
            reason=reason,
            reference_number=reference_number,
            remarks=remarks
        )

        AuditLog.objects.create(
            action='STOCK_ADDED',
            target_type='Item',
            target_id=item.id,
            actioned_by=user.email if hasattr(user, 'email') else 'system',
            comments=f'Manual Stock Addition: {quantity}. Reason: {reason}. Remarks: {remarks}'
        )

        return Response({
            'message': 'Stock added successfully',
            'stock_before': stock_before,
            'stock_after': item.current_stock
        }, status=status.HTTP_200_OK)


class InventoryHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, item_id):
        try:
            item = Item.objects.get(id=item_id)
        except Item.DoesNotExist:
            return Response({'error': 'Item not found'}, status=status.HTTP_404_NOT_FOUND)

        # Order by timestamp ascending for chronological history tracing
        ledgers = StockLedger.objects.filter(item=item).order_by('timestamp')
        history_data = []
        for ledger in ledgers:
            history_data.append({
                'date': ledger.timestamp.isoformat() if ledger.timestamp else None,
                'transaction_type': ledger.transaction_type,
                'quantity': ledger.quantity,
                'stock_before': ledger.stock_before,
                'stock_after': ledger.stock_after,
                'source_type': ledger.source_type or 'SYSTEM',
                'source_id': ledger.source_id or 'OPENING-STOCK'
            })

        return Response({
            'item_id': item.id,
            'item_name': item.name,
            'current_stock': item.current_stock,
            'history': history_data
        }, status=status.HTTP_200_OK)


from django.utils import timezone
from vendors.models import AuditLog

class QCGRNListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        grns = GRN.objects.all().order_by('-created_at')
        
        result = []
        for grn in grns:
            result.append({
                'id': grn.id,
                'grn_number': grn.id,
                'po_id': grn.po_id,
                'po_reference': grn.po_id,
                'vendor_name': grn.vendor_name,
                'received_date': grn.received_date.isoformat() if grn.received_date else '',
                'status': grn.status,
                'inspected_by': grn.inspected_by or '',
                'inspected_at': grn.inspected_at.isoformat() if grn.inspected_at else None,
                'items': grn.items
            })
        return Response(result, status=status.HTTP_200_OK)


class QCGRNDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, grn_id):
        try:
            grn = GRN.objects.get(id=grn_id)
        except GRN.DoesNotExist:
            return Response({'error': 'GRN not found'}, status=status.HTTP_404_NOT_FOUND)
        
        from procurement.models import PurchaseOrder
        from vendors.models import Vendor

        po = PurchaseOrder.objects.filter(id=grn.po_id).first()
        
        vendor_data = {
            "id": "",
            "name": grn.vendor_name or "System Vendor"
        }
        purchase_order_data = None
        attachments = []

        if po:
            vendor_obj = Vendor.objects.filter(id=po.vendor).first()
            if vendor_obj:
                vendor_data = {
                    "id": vendor_obj.id,
                    "name": vendor_obj.name
                }
            else:
                vendor_data = {
                    "id": po.vendor,
                    "name": po.vendor_name
                }
            purchase_order_data = {
                "id": po.id,
                "po_number": po.id,
                "po_date": po.start_date.isoformat() if po.start_date else "",
                "department": po.category or "",
                "requester": po.created_by or "",
                "expected_delivery_date": po.end_date.isoformat() if po.end_date else ""
            }
            if isinstance(po.attachments, list):
                attachments = po.attachments

        formatted_items = []
        for item in grn.items:
            item_name = item.get('item_name') or item.get('itemName') or 'Unknown Item'
            ordered_qty = item.get('ordered_qty') or item.get('orderedQty') or 0
            received_qty = item.get('received_qty') or item.get('receivedQty') or 0
            accepted_qty = item.get('accepted_qty') or item.get('acceptedQty') or 0
            rejected_qty = item.get('rejected_qty') or item.get('rejectedQty') or 0
            remarks = item.get('remarks') or ''
            
            unit_price = 0.0
            if po and isinstance(po.items, list):
                po_item = next((pi for pi in po.items if pi.get('itemName') == item_name or pi.get('itemId') == item.get('item_id') or pi.get('itemId') == item.get('itemId')), None)
                if po_item:
                    unit_price = float(po_item.get('rate') or po_item.get('unit_price') or 0.0)
            
            if not unit_price:
                item_obj = Item.objects.filter(name=item_name).first()
                if item_obj:
                    unit_price = float(item_obj.unit_price)

            formatted_items.append({
                "item_id": item.get('item_id') or item.get('itemId') or '',
                "item_name": item_name,
                "ordered_quantity": ordered_qty,
                "received_quantity": received_qty,
                "accepted_quantity": accepted_qty,
                "rejected_quantity": rejected_qty,
                "unit_price": unit_price,
                "remarks": remarks,
                "uom": item.get('uom') or 'Nos'
            })

        return Response({
            "id": grn.id,
            "grn_number": grn.id,
            "po_reference": grn.po_id,
            "vendor": vendor_data,
            "invoice_number": grn.invoice_number or "",
            "received_date": grn.received_date.isoformat() if grn.received_date else "",
            "received_by": grn.received_by,
            "status": grn.status,
            "inspected_by": grn.inspected_by or "",
            "inspected_at": grn.inspected_at.isoformat() if grn.inspected_at else None,
            "purchase_order": purchase_order_data,
            "items": formatted_items,
            "attachments": attachments
        }, status=status.HTTP_200_OK)


class QCGRNInspectView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, grn_id):
        if request.user.role not in ['store_keeper', 'super_admin']:
            return Response({'error': 'Site Keepers and other non-authorized roles cannot perform inspections.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            grn = GRN.objects.get(id=grn_id)
        except GRN.DoesNotExist:
            return Response({'error': 'GRN not found'}, status=status.HTTP_404_NOT_FOUND)

        if grn.status not in ['pending', 'pending_qc']:
            return Response({'error': 'Inspection can only be performed on pending GRNs.'}, status=status.HTTP_400_BAD_REQUEST)

        items_data = request.data.get('items', [])
        if not items_data:
            return Response({'error': 'No inspection items details provided.'}, status=status.HTTP_400_BAD_REQUEST)

        updated_items = []
        for grn_item in grn.items:
            item_name = grn_item.get('item_name') or grn_item.get('itemName')
            item_id = grn_item.get('item_id') or grn_item.get('itemId')
            
            req_item = next((
                i for i in items_data 
                if i.get('item_name') == item_name or i.get('itemName') == item_name 
                or i.get('item_id') == item_id or i.get('itemId') == item_id
            ), None)

            if not req_item:
                return Response({'error': f'Inspection details missing for item: {item_name}'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                accepted_qty = int(req_item.get('accepted_qty', req_item.get('acceptedQty', 0)))
                rejected_qty = int(req_item.get('rejected_qty', req_item.get('rejectedQty', 0)))
            except ValueError:
                return Response({'error': 'Quantities must be valid integers.'}, status=status.HTTP_400_BAD_REQUEST)

            received_qty = int(grn_item.get('received_qty', grn_item.get('receivedQty', 0)))

            if accepted_qty + rejected_qty != received_qty:
                return Response({
                    'error': f'Accepted Qty ({accepted_qty}) + Rejected Qty ({rejected_qty}) must equal Received Qty ({received_qty}) for {item_name}.'
                }, status=status.HTTP_400_BAD_REQUEST)

            remarks = req_item.get('remarks', '').strip()
            if rejected_qty > 0 and not remarks:
                return Response({
                    'error': f'Remarks/Rejection reason is mandatory for rejected quantities of {item_name}.'
                }, status=status.HTTP_400_BAD_REQUEST)

            grn_item['accepted_qty'] = accepted_qty
            grn_item['rejected_qty'] = rejected_qty
            grn_item['remarks'] = remarks
            grn_item['acceptedQty'] = accepted_qty
            grn_item['rejectedQty'] = rejected_qty

            updated_items.append(grn_item)

        grn.items = updated_items
        grn.status = 'qc_completed'
        grn.inspected_by = request.user.name or request.user.email or 'Store Keeper'
        grn.inspected_at = timezone.now()
        grn.save()

        AuditLog.objects.create(
            action='QC_COMPLETED',
            target_type='GRN',
            target_id=grn.id,
            actioned_by=request.user.email if hasattr(request.user, 'email') else 'system',
            comments=f'Quality inspection completed for GRN {grn.id} by {grn.inspected_by}'
        )

        return Response({'message': 'Inspection completed successfully'}, status=status.HTTP_200_OK)


